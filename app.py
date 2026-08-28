"""
app.py
发票邮件筛选小工具 - Flask 主程序，Zoho 原生搜索版。
部署说明见 README_WEBSITE.md。
"""

import io
import os
import secrets
import threading
import zipfile
from datetime import datetime

from flask import Flask, redirect, url_for, session, request, render_template, jsonify, flash, send_file
from sqlalchemy import inspect, text, func, event
from sqlalchemy.engine import Engine
from dotenv import load_dotenv

load_dotenv()

from models import db, User, ManifestEntry, ProcessedMessage, RunLog, RunStatus, VendorFieldPreset
import oauth_zoho as zoho_oauth
import crypto_util as token_crypto
import ai_extract
import field_config
from mail_sync import run_sync_for_user, DOWNLOADS_ROOT, request_stop, get_run_error_count
from db_utils import commit_with_retry

try:
    from flask_wtf import CSRFProtect
    _CSRF_AVAILABLE = True
except ImportError:
    _CSRF_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


def _admin_emails():
    return {e.strip().lower() for e in os.environ.get("ADMIN_EMAILS", "").split(",") if e.strip()}


def _allowed_login_emails():
    return {e.strip().lower() for e in os.environ.get("ALLOWED_LOGIN_EMAILS", "").split(",") if e.strip()}


def _allowed_login_domains():
    return {d.strip().lower().lstrip("@") for d in os.environ.get("ALLOWED_LOGIN_DOMAINS", "").split(",") if d.strip()}


def _login_allowed(email):
    emails = _allowed_login_emails()
    domains = _allowed_login_domains()
    if not emails and not domains:
        return True
    email_lower = (email or "").lower()
    if email_lower in emails:
        return True
    domain = email_lower.split("@")[-1] if "@" in email_lower else ""
    return domain in domains


def _sqlite_column_ddl(column):
    dialect = db.engine.dialect
    col_type = column.type.compile(dialect=dialect)
    parts = [col_type]
    default = column.default
    if default is not None and getattr(default, "is_scalar", False):
        default_val = default.arg
        if isinstance(default_val, bool):
            if dialect.name == "postgresql":
                parts.append(f"DEFAULT {'TRUE' if default_val else 'FALSE'}")
            else:
                parts.append(f"DEFAULT {1 if default_val else 0}")
        elif isinstance(default_val, (int, float)):
            parts.append(f"DEFAULT {default_val}")
        elif isinstance(default_val, str):
            escaped = default_val.replace("'", "''")
            parts.append(f"DEFAULT '{escaped}'")
    return " ".join(parts)


def _run_lightweight_migrations():
    """没接 Alembic，db.create_all() 只建新表不加新列。这里自动比对 models.py 和数据库的列，
    缺什么自动 ALTER TABLE 补上，升级字段不用手动删库。"""
    inspector = inspect(db.engine)
    for model in (User, ProcessedMessage, ManifestEntry, RunLog, RunStatus):
        table = model.__table__
        if not inspector.has_table(table.name):
            continue
        existing = {c["name"] for c in inspector.get_columns(table.name)}
        missing = [c for c in table.columns if c.name not in existing]
        if not missing:
            continue
        with db.engine.begin() as conn:
            for col in missing:
                ddl = _sqlite_column_ddl(col)
                conn.execute(text(f'ALTER TABLE "{table.name}" ADD COLUMN "{col.name}" {ddl}'))
                print(f"[启动迁移] {table.name} 表补上了缺失的列：{col.name}")


def is_admin_user(user):
    """默认第一个登录的人是管理员；也可用 ADMIN_EMAILS 环境变量（逗号分隔）显式指定。"""
    if not user:
        return False
    configured = _admin_emails()
    if configured:
        return user.zoho_email.lower() in configured
    first_user = User.query.order_by(User.id.asc()).first()
    return bool(first_user and first_user.id == user.id)


_PRESET_SEEDS = [
    {
        "match_pattern": "sentfrom@ascendtms.com",
        "extract_fields": '[{"name": "Date", "aliases": []}, {"name": "Reference", "aliases": []}, {"name": "Amount", "aliases": []}, {"name": "Total", "aliases": []}]',
    },
    {
        "match_pattern": "info@freightcom.com",
        "extract_fields": '[{"name": "Invoice Date", "aliases": []}, {"name": "Amount Paid", "aliases": []}, {"name": "BOL", "aliases": []}, {"name": "Customer Ref", "aliases": []}, {"name": "Charges", "aliases": []}]',
    },
    {
        "match_pattern": "accounts@kwalitylogistics.com",
        "extract_fields": '[{"name": "Date", "aliases": []}, {"name": "Invoice", "aliases": []}, {"name": "Container", "aliases": []}, {"name": "Cust. Ref.#", "aliases": []}, {"name": "Charges", "aliases": []}, {"name": "Grand Total", "aliases": []}]',
    },
]


def _seed_vendor_presets():
    if VendorFieldPreset.query.first():
        return
    for p in _PRESET_SEEDS:
        db.session.add(VendorFieldPreset(**p))
    commit_with_retry(db.session)
    print(f"[seed] 已写入 {len(_PRESET_SEEDS)} 条供应商预设")


def create_app():
    app = Flask(__name__)
    app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

    # 反向代理平台（Render 等）处理 https 后转内部 http，加 ProxyFix 才能正确识别原始协议。
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

    db_url = os.environ.get("DATABASE_URL", "sqlite:///app.db")
    # SQLAlchemy 2.x 需要 postgresql:// 而非 postgres://
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = db_url
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    is_sqlite = db_url.startswith("sqlite")
    if is_sqlite:
        # busy_timeout 防止后台线程写、前端轮询读时 "database is locked"；WAL 让读不等写。
        app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"connect_args": {"timeout": 30}}
    else:
        # Postgres 免费实例连接数有限，pool 调小；pool_pre_ping 防止闲置连接被静默断开。
        app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
            "pool_pre_ping": True,
            "pool_size": 3,
            "max_overflow": 2,
        }

    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
    # 本机 http 测试时不能开 Secure；正式部署 https 时在 .env 里设 FORCE_HTTPS_COOKIES=1。
    app.config["SESSION_COOKIE_SECURE"] = os.environ.get("FORCE_HTTPS_COOKIES", "0") == "1"

    db.init_app(app)

    if is_sqlite:
        @event.listens_for(Engine, "connect")
        def _sqlite_pragma(dbapi_connection, connection_record):
            cursor = dbapi_connection.cursor()
            cursor.execute("PRAGMA journal_mode=WAL")
            cursor.execute("PRAGMA busy_timeout=30000")
            cursor.close()

    if _CSRF_AVAILABLE:
        CSRFProtect(app)
    else:
        print("[警告] 没装 Flask-WTF，CSRF 防护未启用。建议执行 pip install -r requirements.txt 后重启。")
        app.jinja_env.globals.setdefault("csrf_token", lambda: "")

    with app.app_context():
        db.create_all()
        _run_lightweight_migrations()
        _seed_vendor_presets()
        # 清理服务器重启前残留的 is_running=True 状态
        stale = RunStatus.query.filter_by(is_running=True).all()
        if stale:
            for s in stale:
                s.is_running = False
                s.stop_requested = False
            commit_with_retry(db.session)

    register_routes(app)
    return app


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return User.query.get(uid)


def _wants_json():
    """前端 AJAX 请求带此头，后端据此返回 JSON 而非 flash+redirect。"""
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def _respond(message, ok=True, **extra):
    if _wants_json():
        payload = {"ok": ok, "message": message}
        payload.update(extra)
        return jsonify(payload)
    flash(message)
    return redirect(url_for("dashboard"))


RECORDS_PAGE_SIZE = 30


def _query_manifest_page(user_id, page):
    """处理记录分页查询，被 dashboard、/manifest AJAX 接口、Excel 导出共用。"""
    page = page if page and page >= 1 else 1
    q = ManifestEntry.query.filter_by(user_id=user_id).order_by(
        ManifestEntry.created_at.desc(), ManifestEntry.id.desc()
    )
    total = q.count()
    total_pages = max(1, (total + RECORDS_PAGE_SIZE - 1) // RECORDS_PAGE_SIZE)
    page = min(page, total_pages)
    rows = q.offset((page - 1) * RECORDS_PAGE_SIZE).limit(RECORDS_PAGE_SIZE).all()
    return rows, page, total_pages, total


def _all_extract_field_names(user):
    """全局字段 + 所有 vendor 预设字段的并集，用于表格列和 Excel 导出。
    全局字段在前，预设独有字段按配置顺序追加，同名字段只出现一次。"""
    names = list(field_config.field_names_only(user.extract_fields))
    for p in VendorFieldPreset.query.order_by(VendorFieldPreset.id.asc()).all():
        for name in field_config.field_names_only(p.extract_fields):
            if name not in names:
                names.append(name)
    return names


def _active_field_names(user):
    """在 _all_extract_field_names 的基础上，只保留该用户数据中至少有一行有值的字段列。
    过滤后为空（如 AI 全部失败）时退回完整列表，不让表格连一个字段列都没有。"""
    all_names = _all_extract_field_names(user)
    if not all_names:
        return all_names
    all_names_set = set(all_names)
    # extracted_fields 是 @property，不能传给 with_entities，用底层 DB Column。
    import json as _json
    rows = (
        ManifestEntry.query
        .filter_by(user_id=user.id)
        .with_entities(ManifestEntry.extracted_fields_json)
        .all()
    )
    cols_with_data = {
        name
        for (json_str,) in rows
        if json_str
        for name, val in (_json.loads(json_str) if isinstance(json_str, str) else {}).items()
        if val and name in all_names_set
    }
    return [n for n in all_names if n in cols_with_data] or all_names


def _pick_field_names_for_member(user, sender_raw):
    """普通用户/预览成员视角下，按发件人值挑选字段列名：
    comma-split 后对每个 token 都尝试匹配 vendor 预设，把所有命中预设的字段取并集；
    全部 token 都没命中时，退回全局默认字段。"""
    global_defs = field_config.parse_extract_fields(user.extract_fields)
    all_presets = VendorFieldPreset.query.order_by(VendorFieldPreset.id.asc()).all()
    matched_names: list[str] = []
    seen: set[str] = set()
    if all_presets:
        presets_data = [(p.match_pattern, field_config.parse_extract_fields(p.extract_fields)) for p in all_presets]
        for token in [t.strip() for t in (sender_raw or "").split(",") if t.strip()]:
            chosen = field_config.pick_field_defs_for_sender(token, global_defs, presets_data)
            if chosen is not global_defs:
                for f in chosen:
                    if f["name"] not in seen:
                        seen.add(f["name"])
                        matched_names.append(f["name"])
    if matched_names:
        return matched_names
    return field_config.field_names_only(user.extract_fields)


def _reset_processing_state(user_id):
    """运行前清空去重记录和处理记录表格，每次都全量重新扫描。"""
    ProcessedMessage.query.filter_by(user_id=user_id).delete()
    ManifestEntry.query.filter_by(user_id=user_id).delete()
    commit_with_retry(db.session)


def register_routes(app):
    @app.route("/")
    def index():
        if current_user():
            return redirect(url_for("dashboard"))
        return render_template("login.html")

    # ---------------- Zoho 登录 ----------------

    @app.route("/login/zoho")
    def login_zoho():
        state = secrets.token_hex(16)
        session["zoho_oauth_state"] = state
        try:
            return redirect(zoho_oauth.build_authorize_url(state))
        except RuntimeError as e:
            flash(str(e))
            return redirect(url_for("index"))

    @app.route("/auth/zoho/callback")
    def auth_zoho_callback():
        error = request.args.get("error")
        if error:
            flash(f"Zoho 授权失败：{error}")
            return redirect(url_for("index"))

        state = request.args.get("state")
        if not state or state != session.get("zoho_oauth_state"):
            flash("授权状态校验失败，请重新登录。")
            return redirect(url_for("index"))

        code = request.args.get("code")
        accounts_server = request.args.get("accounts-server", zoho_oauth.DEFAULT_ACCOUNTS_SERVER)

        try:
            token_resp = zoho_oauth.exchange_code_for_token(code, accounts_server)
            access_token = token_resp["access_token"]
            refresh_token = token_resp.get("refresh_token")
            api_domain = zoho_oauth.derive_mail_api_domain(accounts_server)
            email_addr, account_id = zoho_oauth.get_account_info(access_token, api_domain)
        except Exception as e:
            flash(f"Zoho 授权处理失败：{e}")
            return redirect(url_for("index"))

        if not _login_allowed(email_addr):
            flash(f"{email_addr} 不在允许登录的名单里，如果这是误拦请联系管理员。")
            return redirect(url_for("index"))

        user = User.query.filter_by(zoho_email=email_addr).first()
        if not user:
            user = User(zoho_email=email_addr)
            db.session.add(user)

        if refresh_token:
            user.zoho_refresh_token = token_crypto.encrypt(refresh_token)
        user.zoho_accounts_server = accounts_server
        user.zoho_api_domain = api_domain
        user.zoho_account_id = account_id
        user.last_login_at = datetime.utcnow()
        user.login_count = (user.login_count or 0) + 1
        commit_with_retry(db.session)

        session["user_id"] = user.id
        return redirect(url_for("dashboard"))

    @app.route("/logout")
    def logout():
        session.clear()
        return redirect(url_for("index"))

    # ---------------- 仪表盘 / 配置 / 运行 ----------------

    @app.route("/dashboard")
    def dashboard():
        user = current_user()
        if not user:
            return redirect(url_for("index"))

        page = request.args.get("page", 1, type=int) or 1
        rows, page, total_pages, total_entries = _query_manifest_page(user.id, page)

        # 管理员加 ?view=member 可预览普通用户视角，不影响其他接口的权限判断。
        real_is_admin = is_admin_user(user)
        previewing_member = real_is_admin and request.args.get("view") == "member"
        effective_is_admin = real_is_admin and not previewing_member

        status = RunStatus.query.get(user.id)
        return render_template(
            "dashboard.html",
            user=user,
            entries=rows,
            status=status,
            is_admin=effective_is_admin,
            real_is_admin=real_is_admin,
            previewing_member=previewing_member,
            page=page,
            total_pages=total_pages,
            total_entries=total_entries,
            # 管理员视角显示所有预设字段列；普通用户按已保存的发件人过滤值挑预设。
            extract_field_names=_active_field_names(user) if effective_is_admin else _pick_field_names_for_member(user, user.search_sender_contains),
            extract_field_defs=field_config.parse_extract_fields(user.extract_fields),
            # vendor_presets 完整内容只给管理员用于渲染/编辑。
            # vendor_preset_patterns 给所有用户做"发件人命中提示"。
            # vendor_preset_field_names 给普通用户前端动态切换表格列用（pattern → 字段名列表）。
            vendor_presets=[
                {
                    "id": p.id,
                    "match_pattern": p.match_pattern,
                    "field_defs": field_config.parse_extract_fields(p.extract_fields),
                }
                for p in VendorFieldPreset.query.order_by(VendorFieldPreset.id.asc()).all()
            ] if effective_is_admin else [],
            vendor_preset_patterns=[
                p.match_pattern
                for p in VendorFieldPreset.query.order_by(VendorFieldPreset.id.asc()).all()
            ],
            vendor_preset_field_names=[
                {"match_pattern": p.match_pattern, "field_names": field_config.field_names_only(p.extract_fields)}
                for p in VendorFieldPreset.query.order_by(VendorFieldPreset.id.asc()).all()
            ] if not effective_is_admin else [],
            ai_problem=ai_extract.diagnose(),
        )

    @app.route("/admin/users")
    def admin_users():
        user = current_user()
        if not user:
            return redirect(url_for("index"))
        if not is_admin_user(user):
            flash("这个页面只有管理员能看。")
            return redirect(url_for("dashboard"))

        rows = (
            db.session.query(
                User,
                func.count(ManifestEntry.id).label("saved_count"),
            )
            .outerjoin(ManifestEntry, ManifestEntry.user_id == User.id)
            .group_by(User.id)
            .order_by(User.created_at.asc())
            .all()
        )

        return render_template(
            "admin_users.html",
            user=user,
            rows=rows,
            total_users=len(rows),
        )

    @app.route("/manifest")
    def manifest_page():
        """处理记录表格的 AJAX 刷新接口，运行结束后前端用此更新表格。"""
        user = current_user()
        if not user:
            return jsonify({"error": "not logged in"}), 401

        page = request.args.get("page", 1, type=int) or 1
        rows, page, total_pages, total_entries = _query_manifest_page(user.id, page)
        real_is_admin = is_admin_user(user)
        previewing_member = real_is_admin and request.args.get("view") == "member"
        effective_is_admin = real_is_admin and not previewing_member
        if effective_is_admin:
            field_names = _active_field_names(user)
        else:
            # 前端把当前"发件人包含"输入框的值传过来，据此挑对应预设的字段列
            sender = request.args.get("sender", user.search_sender_contains or "")
            field_names = _pick_field_names_for_member(user, sender)

        entries_json = [
            {
                "sender_name": e.sender_name,
                "sender_email": e.sender_email,
                "subject": e.subject,
                "filename": e.original_filename,
                # 重复行复用"正主"那一行的下载链接。
                "download_url": (
                    url_for("download_attachment", entry_id=e.duplicate_of_id)
                    if e.is_duplicate and e.duplicate_of_id
                    else url_for("download_attachment", entry_id=e.id) if e.saved_filename else None
                ),
                "is_duplicate": bool(e.is_duplicate),
                "fields": [e.extracted_fields.get(name) for name in field_names],
            }
            for e in rows
        ]

        return jsonify(
            {
                "entries": entries_json,
                "field_names": field_names,
                "page": page,
                "total_pages": total_pages,
                "total_entries": total_entries,
            }
        )

    @app.route("/config", methods=["POST"])
    def save_config():
        user = current_user()
        if not user:
            return redirect(url_for("index"))

        user.search_subject_contains = request.form.get("search_subject_contains", "")
        user.search_attachment_contains = request.form.get("search_attachment_contains", "")
        user.search_sender_contains = request.form.get("search_sender_contains", "")
        user.search_content_contains = request.form.get("search_content_contains", "")
        user.search_sender_excludes = request.form.get("search_sender_excludes", "")
        user.search_since_date = request.form.get("search_since_date", "").strip()
        user.search_until_date = request.form.get("search_until_date", "").strip()
        user.search_require_attachment = bool(request.form.get("search_require_attachment"))
        # 先解析再重新序列化，顺手清洗空字段名/空备选名，同时兼容老格式。
        user.extract_fields = field_config.serialize_extract_fields(
            field_config.parse_extract_fields(request.form.get("extract_fields", ""))
        )
        user.ai_extract_enabled = bool(request.form.get("ai_extract_enabled"))

        commit_with_retry(db.session)
        return _respond("配置已保存。")

    # ---------------- 按发件人配置的 AI 提取字段预设（全站共享，仅管理员可维护） ----------------

    @app.route("/vendor-presets", methods=["POST"])
    def create_vendor_preset():
        user = current_user()
        if not user:
            return redirect(url_for("index"))
        if not is_admin_user(user):
            return _respond("这个预设是全站共享的配置，只有管理员能改。", ok=False)

        match_pattern = request.form.get("match_pattern", "").strip()
        if not match_pattern:
            return _respond("请先填发件人邮箱或域名，再添加预设。", ok=False)

        preset = VendorFieldPreset(
            match_pattern=match_pattern,
            extract_fields=field_config.serialize_extract_fields(
                field_config.parse_extract_fields(request.form.get("extract_fields", ""))
            ),
        )
        db.session.add(preset)
        commit_with_retry(db.session)
        return _respond("预设已添加。", id=preset.id)

    @app.route("/vendor-presets/<int:preset_id>", methods=["POST"])
    def update_vendor_preset(preset_id):
        user = current_user()
        if not user:
            return redirect(url_for("index"))
        if not is_admin_user(user):
            return _respond("这个预设是全站共享的配置，只有管理员能改。", ok=False)

        preset = VendorFieldPreset.query.filter_by(id=preset_id).first()
        if not preset:
            return _respond("没找到这条预设（可能已经在别的地方被删除了，刷新一下页面看看）。", ok=False)

        match_pattern = request.form.get("match_pattern", "").strip()
        if not match_pattern:
            return _respond("发件人邮箱/域名不能留空。", ok=False)

        preset.match_pattern = match_pattern
        preset.extract_fields = field_config.serialize_extract_fields(
            field_config.parse_extract_fields(request.form.get("extract_fields", ""))
        )
        commit_with_retry(db.session)
        return _respond("预设已保存。")

    @app.route("/vendor-presets/<int:preset_id>/delete", methods=["POST"])
    def delete_vendor_preset(preset_id):
        user = current_user()
        if not user:
            return redirect(url_for("index"))
        if not is_admin_user(user):
            return _respond("这个预设是全站共享的配置，只有管理员能改。", ok=False)

        preset = VendorFieldPreset.query.filter_by(id=preset_id).first()
        if preset:
            db.session.delete(preset)
            commit_with_retry(db.session)
        return _respond("预设已删除。")

    @app.route("/run", methods=["POST"])
    def run_now():
        user = current_user()
        if not user:
            return redirect(url_for("index"))

        status = RunStatus.query.get(user.id)
        if status and status.is_running:
            return _respond("已经有一个任务在运行了，请等它跑完。", ok=False)

        _reset_processing_state(user.id)

        app = current_app_ref["app"]
        thread = threading.Thread(target=run_sync_for_user, args=(app, user.id), daemon=True)
        thread.start()

        return _respond("已清空之前的处理记录，开始全量重新扫描，下面的日志会自动刷新。")

    @app.route("/run_titles_only", methods=["POST"])
    def run_titles_only():
        """仅导出标题模式：跳过附件下载，只记录邮件标题/发件人/日期，速度更快。"""
        user = current_user()
        if not user:
            return redirect(url_for("index"))

        status = RunStatus.query.get(user.id)
        if status and status.is_running:
            return _respond("已经有一个任务在运行了，请等它跑完。", ok=False)

        _reset_processing_state(user.id)

        app = current_app_ref["app"]
        thread = threading.Thread(
            target=run_sync_for_user, args=(app, user.id), kwargs={"download_attachments": False}, daemon=True
        )
        thread.start()

        return _respond("已清空之前的处理记录，开始全量重新扫描（仅导出标题模式，不下载附件），下面的日志会自动刷新。")

    @app.route("/download_attachment/<int:entry_id>")
    def download_attachment(entry_id):
        user = current_user()
        if not user:
            return redirect(url_for("index"))

        entry = ManifestEntry.query.filter_by(id=entry_id, user_id=user.id).first()
        if not entry:
            flash("找不到这个附件记录，可能已经被清空了。")
            return redirect(url_for("dashboard"))

        folder = os.path.join(DOWNLOADS_ROOT, str(user.id), entry.run_id or "")
        filepath = os.path.join(folder, entry.saved_filename or "")
        if not entry.saved_filename or not os.path.isfile(filepath):
            flash("这个附件的文件已经不在服务器上了（临时文件可能已被清理），建议重新运行一次。")
            return redirect(url_for("dashboard"))

        return send_file(filepath, as_attachment=True, download_name=entry.original_filename or "attachment")

    @app.route("/download/<run_id>")
    def download_run(run_id):
        """把某次运行的附件打包成 ZIP 下载。"""
        user = current_user()
        if not user:
            return redirect(url_for("index"))

        exists = ManifestEntry.query.filter_by(user_id=user.id, run_id=run_id).first()
        if not exists:
            flash("找不到这次运行的记录。")
            return redirect(url_for("dashboard"))

        folder = os.path.join(DOWNLOADS_ROOT, str(user.id), run_id)
        if not os.path.isdir(folder) or not os.listdir(folder):
            flash("这次运行的文件已经不在服务器上了（临时文件可能已被清理），建议重新运行一次。")
            return redirect(url_for("dashboard"))

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for name in os.listdir(folder):
                fpath = os.path.join(folder, name)
                if os.path.isfile(fpath):
                    zf.write(fpath, arcname=name)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/zip",
            as_attachment=True,
            download_name=f"invoices_{run_id}.zip",
        )

    @app.route("/export_excel")
    def export_excel():
        """导出处理记录为 Excel：按发件人匹配的供应商预设分 sheet，
        每个 sheet 只含该预设的字段列；无匹配预设的记录放「默认字段」sheet。"""
        import re as _re
        user = current_user()
        if not user:
            return redirect(url_for("index"))
        if not _OPENPYXL_AVAILABLE:
            flash("服务器还没装 openpyxl，导出不了 Excel。请执行 pip install -r requirements.txt 后重启。")
            return redirect(url_for("dashboard"))

        rows = (
            ManifestEntry.query.filter_by(user_id=user.id)
            .order_by(ManifestEntry.created_at.desc(), ManifestEntry.id.desc())
            .all()
        )

        global_defs = field_config.parse_extract_fields(user.extract_fields)
        global_field_names = field_config.field_names_only(user.extract_fields)
        all_presets = VendorFieldPreset.query.order_by(VendorFieldPreset.id.asc()).all()
        presets_data = [
            (p.match_pattern, field_config.parse_extract_fields(p.extract_fields))
            for p in all_presets
        ]

        # 按预设分组：{match_pattern: [rows]}，未命中的放 default_rows
        from collections import OrderedDict
        groups: OrderedDict[str, list] = OrderedDict()
        pattern_fields: dict[str, list[str]] = {}
        default_rows: list = []

        for e in rows:
            sender = (e.sender_email or "").strip()
            assigned = False
            for pattern, field_defs in presets_data:
                chosen = field_config.pick_field_defs_for_sender(sender, global_defs, [(pattern, field_defs)])
                if chosen is not global_defs:
                    if pattern not in groups:
                        groups[pattern] = []
                        pattern_fields[pattern] = [f["name"] for f in field_defs]
                    groups[pattern].append(e)
                    assigned = True
                    break
            if not assigned:
                default_rows.append(e)

        fixed_headers = ["发件人名", "发件人邮箱", "主题", "附件标题", "下载链接", "备注"]
        fixed_widths = [22, 28, 34, 34, 46, 34]
        link_col_idx = fixed_headers.index("下载链接") + 1

        def _write_sheet(ws, sheet_rows, field_names):
            # 只保留有数据的字段列
            cols_with_data = {
                name for r in sheet_rows
                for name, val in r.extracted_fields.items()
                if val and name in field_names
            } if sheet_rows and field_names else set()
            active_fields = [n for n in field_names if n in cols_with_data] or list(field_names)
            ws.append(fixed_headers + [f"{n}(AI提取)" for n in active_fields])
            for e in sheet_rows:
                link_entry_id = e.duplicate_of_id if (e.is_duplicate and e.duplicate_of_id) else e.id
                has_file = bool(e.saved_filename) or bool(e.is_duplicate and e.duplicate_of_id)
                link = url_for("download_attachment", entry_id=link_entry_id, _external=True) if has_file else ""
                note = "重复（内容与另一封邮件相同，已保留发送时间更晚的那份）" if e.is_duplicate else ""
                ws.append(
                    [e.sender_name or "", e.sender_email or "", e.subject or "",
                     e.original_filename or "", link, note]
                    + [e.extracted_fields.get(n) or "" for n in active_fields]
                )
                if link:
                    cell = ws.cell(row=ws.max_row, column=link_col_idx)
                    cell.hyperlink = link
                    cell.style = "Hyperlink"
            for i, w in enumerate(fixed_widths + [18] * len(active_fields), start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def _safe_title(name, idx):
            safe = _re.sub(r'[\\/?*\[\]:]', '_', name)
            return (safe[:28] + f"_{idx}" if len(safe) > 31 else safe) or f"Sheet{idx}"

        wb = Workbook()
        wb.remove(wb.active)

        for i, (pattern, pattern_rows) in enumerate(groups.items(), start=1):
            ws = wb.create_sheet(title=_safe_title(pattern, i))
            _write_sheet(ws, pattern_rows, pattern_fields[pattern])

        if default_rows:
            ws = wb.create_sheet(title="默认字段")
            _write_sheet(ws, default_rows, list(global_field_names))

        if not wb.worksheets:
            wb.create_sheet(title="处理记录")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="处理记录.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/stop", methods=["POST"])
    def stop_run():
        user = current_user()
        if not user:
            if _wants_json():
                return jsonify({"ok": False, "message": "未登录，请刷新页面重新登录。"})
            return redirect(url_for("index"))

        status = RunStatus.query.get(user.id)
        if not (status and status.is_running):
            return _respond("当前没有正在运行的任务。", ok=False)

        # 先在内存里设停止标志（不涉及 DB，立刻生效），再尝试写库（兜底，可失败）。
        request_stop(user.id)
        try:
            status.stop_requested = True
            db.session.commit()
        except Exception:
            db.session.rollback()

        return _respond("已发送停止请求，正在停止当前运行 ...")

    @app.route("/clear_processed", methods=["POST"])
    def clear_processed():
        user = current_user()
        if not user:
            return redirect(url_for("index"))

        status = RunStatus.query.get(user.id)
        if status and status.is_running:
            return _respond("有任务正在运行，请先停止或等它跑完再清空处理记录。", ok=False)

        deleted = ProcessedMessage.query.filter_by(user_id=user.id).delete()
        commit_with_retry(db.session)
        return _respond(f"已清空处理记录（{deleted} 条）。下次运行会重新扫描所有历史邮件，已经保存过的附件可能会重复保存一份。")

    @app.route("/clear_manifest", methods=["POST"])
    def clear_manifest():
        """清空处理记录表格（ManifestEntry），下载链接随之失效。"""
        user = current_user()
        if not user:
            return redirect(url_for("index"))

        status = RunStatus.query.get(user.id)
        if status and status.is_running:
            return _respond("有任务正在运行，请先停止或等它跑完再清空附件历史。", ok=False)

        deleted = ManifestEntry.query.filter_by(user_id=user.id).delete()
        commit_with_retry(db.session)
        return _respond(f"已清空附件历史（{deleted} 条）。之前的下载链接会跟着失效，不影响下次是否重新扫描邮件。")

    @app.route("/run/status")
    def run_status():
        user = current_user()
        if not user:
            return jsonify({"error": "not logged in"}), 401

        status = RunStatus.query.get(user.id)
        is_running = bool(status and status.is_running)
        run_id = status.current_run_id if status else None

        # 详细日志只下发给管理员，避免前端开发者工具里泄露技术细节。
        logs = []
        if run_id and is_admin_user(user):
            rows = (
                RunLog.query.filter_by(user_id=user.id, run_id=run_id)
                .order_by(RunLog.created_at.asc())
                .limit(500)
                .all()
            )
            logs = [r.message for r in rows]

        download_url = None
        if run_id and (status.saved_count if status else 0):
            download_url = url_for("download_run", run_id=run_id)

        error_count = get_run_error_count(user.id, run_id) if run_id else 0

        payload = {
            "is_running": is_running,
            "logs": logs,
            "checked": status.checked_count if status else 0,
            "total": status.total_count if status else 0,
            "last_run_ok": bool(status.last_run_ok) if status and status.last_run_ok is not None else True,
            "download_url": download_url,
            "error_count": error_count,
        }
        # 命中/已下载细分数字只给管理员。
        if is_admin_user(user):
            payload["matched"] = status.matched_count if status else 0
            payload["saved"] = status.saved_count if status else 0
        return jsonify(payload)

    @app.route("/api/sender-suggestions")
    def sender_suggestions():
        user = current_user()
        if not user:
            return jsonify([]), 401
        q = (request.args.get("q") or "").strip().lower()
        # 查当前用户所有处理记录里的发件人邮箱，按出现次数降序
        rows = (
            db.session.query(ManifestEntry.sender_email, func.count().label("cnt"))
            .filter(ManifestEntry.user_id == user.id, ManifestEntry.sender_email.isnot(None))
            .group_by(ManifestEntry.sender_email)
            .order_by(func.count().desc())
            .limit(200)
            .all()
        )
        # 构建：完整邮箱 + 域名，都按频率累加
        from collections import defaultdict
        score = defaultdict(int)
        for email, cnt in rows:
            email = email.strip().lower()
            if not email:
                continue
            score[email] += cnt
            if "@" in email:
                domain = email.split("@", 1)[1]
                score[domain] += cnt

        # 把管理员配的供应商预设 match_pattern 也加入候选（给一个基础分，保证没跑过也能推荐）
        for preset in VendorFieldPreset.query.all():
            pat = (preset.match_pattern or "").strip().lower()
            if pat:
                score.setdefault(pat, 0)
                score[pat] = max(score[pat], 1)

        # 过滤匹配 q 的项，按分数排序，最多返回 10 条
        results = sorted(
            [(k, v) for k, v in score.items() if not q or q in k],
            key=lambda x: -x[1]
        )[:10]
        return jsonify([r[0] for r in results])


# 存 app 引用供后台线程使用，避免跨线程使用 current_app 代理对象。
current_app_ref = {}

app = create_app()
current_app_ref["app"] = app


if __name__ == "__main__":
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    host = os.environ.get("HOST", "127.0.0.1")
    app.run(host=host, port=int(os.environ.get("PORT", 5000)), debug=debug_mode, threaded=True)
